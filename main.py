# 屎山代码，懒得优化了，能跑就行。很多地方写的纯纯是一坨屎。

import requests as req
import re
import time
import openpyxl
import random
from random import randint
from colorama import Fore, Back, Style, init

class apf:
    def __init__(self):
        super().__init__()
        self.exit = 1

        init(autoreset=True)

        #你的夸克登录Cookie
        self.cookie = "Your Cookie"
        #获取全部文件的URL
        self.getAllFiles_Url = "https://drive-pc.quark.cn/1/clouddrive/file/sort?pr=ucpro&fr=pc&uc_param_str=&pdir_fid=0&_page=1&_size=100&_fetch_total=false&_fetch_sub_dirs=1&_sort=&__dt=108496&__t=1729840382463"

        self.UserAgent = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ]
        self.random_agent = self.UserAgent[randint(0, len(self.UserAgent) - 1)]
        #get请求的Headers
        self.get_headers = {
            "user-agent": self.random_agent,
            "cookie": self.cookie
        }

        #post请求的headers
        self.post_headers = {
            "user-agent": self.random_agent,
            "cookie": self.cookie,
            'Content-Type': 'application/json'
        }

    def get_merged_cell_value(self,sheet, row, col):
        cell = sheet.cell(row=row, column=col)
        # 检查当前单元格是否是合并单元格的一部分
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 由于我们只关心行合并，检查合并单元格是否跨越多行
                if merged_range.min_col == merged_range.max_col:
                    # 返回合并单元格左上角的值
                    merged_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if merged_cell.value is not None:  # 检查合并单元格的值是否为None
                        return str(merged_cell.value)  # 转换为字符串
        # 如果不是合并单元格的一部分，或者不是行合并，直接返回当前单元格的值
        if cell.value is not None:  # 检查单元格的值是否为None
            return str(cell.value)  # 转换为字符串

    def getQuarkLink(self):
        #读取xlsx表格中的游戏名称跟链接，可以自己修改路径
        workbook = openpyxl.load_workbook('C:/Users/Administrator/Desktop/自行将此文件后面改成表格后缀（10.18）.xlsx')
        sheet = workbook.active  # 或者通过名称获取

        #page_start = 43
        page_start = 931
        page_end = 949
        # page_end = 498
        #page_end = 927
        self.quark_name_List = []
        self.quark_link_List = []
        n = 1
        for row_index in range(page_start, page_end + 1):
            row_values = []
            row_values2 = []
            for col_index in range(2, 3):
                value = self.get_merged_cell_value(sheet, row_index, col_index)
                if value is not None:
                    row_values.append(value)

            for col_index in range(12, 13):  # 假设读取第A到第M列
                value2 = self.get_merged_cell_value(sheet, row_index, col_index)
                if value2 is not None:  # 只添加非None的值
                    row_values2.append(value2)

            if row_values:
                print(str(n) + "." + str(row_values))
                reTxt = re.sub(r'[:<>|*?,/%]', '', str(row_values))
                self.quark_name_List.append(reTxt)
            if row_values2:
                print(str(n) + "." + str(row_values2))
                self.quark_link_List.append(row_values2)
                n += 1
        workbook.close()

    def getUser_AllFiles(self):
        self.allFiles_dict_name = {}
        self.allFiles_dict_Id = {}
        self.allFiles_dict_platformSource = {}
        resu = req.get(url=self.getAllFiles_Url, headers=self.get_headers).text
        takeFile_Name = re.findall(r'"file_name":"([^"]+)"', resu)
        takeFile_Id = re.findall(r'"fid":"([^"]+)"',resu)
        takeFile_platformSource = re.findall(r'"platform_source":"([^"]+)"', resu)
        count = 0
        for name, Id,platformSource in zip(takeFile_Name, takeFile_Id,takeFile_platformSource):
            self.allFiles_dict_name[str(count)] = name
            self.allFiles_dict_Id[str(count)] = Id
            self.allFiles_dict_platformSource[str(count)] = platformSource

            print(Fore.CYAN +str(count) + "." + name)
            count += 1

    def getSelectorFile_Id(self):
        self.selectorFile_Id = input(Fore.RED + "请输入指定文件夹用于保存文件（序号）(退出：-1)：")
        if self.selectorFile_Id != "-1":
            print("当前选择文件夹：" + Fore.YELLOW + self.allFiles_dict_name[self.selectorFile_Id])
            print(Fore.LIGHTBLACK_EX + "{\n\t" +

                  "Id:" + self.allFiles_dict_Id[self.selectorFile_Id] +
                  "\n\tplatform_source:" + self.allFiles_dict_platformSource[self.selectorFile_Id] +

                  "\n}\n")

            self.controll_CreateFiles()
        else:
            fun.exit = 0
            return

    def controll_CreateFiles(self):
        for quark_name,quark_link in zip(self.quark_name_List,self.quark_link_List):
            local_url = "https://drive-pc.quark.cn/1/clouddrive/file?pr=ucpro&fr=pc&uc_param_str="
            data = {
                'pdir_fid': self.allFiles_dict_Id[self.selectorFile_Id],
                'file_name': str(quark_name).replace('[','').replace(']','').replace("'",''),
                'dir_path': "",
                'dir_init_lock': 'false'
            }

            resu = req.post(url=local_url, headers=self.post_headers, json=data)
            print(Fore.LIGHTBLACK_EX + "返回代码：{\n" + resu.text + "\n}\n")
            tempTxt = str(re.findall(r'"status":([^"]+),', resu.text))
            tempTxt = tempTxt.replace('[', '').replace(']', '').replace("'", '')
            if tempTxt == "200":
                print(Fore.LIGHTMAGENTA_EX + Fore.LIGHTMAGENTA_EX + str(quark_name).replace('[','').replace(']','').replace("'",'') + "创建文件夹成功!\n")
                self.quark_links = quark_link
                self.create_file_fid = str(re.search(r'"fid":"([^"]+)"', resu.text).group(1))
                if self.User_getStoken() == 0:
                    break
                time.sleep(random.randint(3, 5))
            else:
                errortxt = str(re.search(r'"message":"([^"]+)"',resu.text).group(1))
                print(Fore.LIGHTMAGENTA_EX + str(quark_name).replace('[','').replace(']','').replace("'",'') +" 文件夹创建错误！错误代码："+errortxt +"\n")

    def User_getStoken(self):
        # 获取stoken
        # https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc&uc_param_str=
        # 获取文件ID
        # https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail?pr=ucpro&fr=pc&uc_param_str=&pwd_id=文件链接的ID&stoken=Stoken令牌&pdir_fid=0&force=0&_page=1&_size=50&_fetch_banner=1&_fetch_share=1&_fetch_total=1&_sort=file_type:asc,updated_at:desc
        print(Fore.LIGHTBLACK_EX + "正在获取Stoken令牌...\n")
        urltext = "https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc&uc_param_str="

        self.fileId = str(self.quark_links).split('/')[-1].replace("'","").replace("]",'')

        post_data = {
            "passcode":"",
            "pwd_id":self.fileId
        }
        print(post_data)
        resu = req.post(url=urltext,headers=self.post_headers, json=post_data).text
        self.stoken = str(re.search('"stoken":"([^"]+)"',resu).group(1))
        if str(re.search(r'"status":([^"]+),',resu).group(1)) == "200":
            print(Fore.LIGHTMAGENTA_EX + "Stoken令牌获取成功！\n返回代码：{\n"+resu+"\n}\n")
            self.User_getFileID()
        else:
            errortxt = str(re.search(r'"message":"([^"]+)"', resu).group(1))
            print(Fore.LIGHTMAGENTA_EX + "Stoken令牌获取失败！" + "错误代码：{\n"+errortxt+"\n}\n")

    def User_getFileID(self):
        url = "https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail?"
        params = {
            'pr': 'ucpro',
            'fr': 'pc',
            'uc_param_str': '',  # 如果这个参数没有值，可以传递空字符串或者不包含这个参数
            'pwd_id': self.fileId,
            'stoken': self.stoken,
            'pdir_fid': '0',
            'force': '0',
            '_page': '1',
            '_size': '50',
            '_fetch_banner': '1',
            '_fetch_share': '1',
            '_fetch_total': '1',
            '_sort': 'file_type:asc,updated_at:desc'
        }

        resu = req.get(url=url, headers=self.get_headers,params=params).text
        status = re.search('"status":([^"]+),',resu).group(1)


        if status == "200":
            self.file_fids = re.search('"first_fid":"([^"]+)"',resu).group(1)
            self.file_share_id = re.search('"share_fid_token":"([^"]+)"',resu).group(1)
            print(Fore.LIGHTMAGENTA_EX + "Fid | shared_id 获取成功！")
            print(Fore.LIGHTBLACK_EX + "fid：{" + self.file_fids + "}")
            print(Fore.LIGHTBLACK_EX + "share：{" + self.file_share_id + "}")

            self.DepositFile()
        else:
            print(Fore.LIGHTMAGENTA_EX + "Fid | share_fid_token 获取失败...\n")
            return

    def DepositFile(self):
        url = "https://drive-pc.quark.cn/1/clouddrive/share/sharepage/save?pr=ucpro&fr=pc&uc_param_str="
        data = {
            "fid_list": [self.file_fids],
            "fid_token_list": [self.file_share_id],
            "pdir_fid": "0",
            "pwd_id": self.fileId,
            "scene":"link",
            "stoken": self.stoken,
            "to_pdir_fid": self.create_file_fid
        }
        #print(data)
        resu = req.post(url=url, headers=self.post_headers, json=data).text
        print(resu)



if __name__ == "__main__":
    fun = apf()

    fun.getQuarkLink()
    fun.getUser_AllFiles()
    fun.getSelectorFile_Id()